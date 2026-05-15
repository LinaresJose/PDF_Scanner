"""
excel_exporter.py — Exporta el DataFrame de resultados a un archivo .xlsx.

Características:
  - Cabecera estilizada con color y negrita.
  - Ancho de columnas ajustado automáticamente al contenido.
  - Hoja de resumen con estadísticas básicas.
  - Filas con error marcadas en rojo.
  - Filtros automáticos habilitados.
"""

import logging
from pathlib import Path

import pandas as pd

logger = logging.getLogger(__name__)


# Mapeo de nombres de columna internos → etiquetas legibles en el Excel
COLUMNAS_LEGIBLES = {
    "archivo": "Archivo PDF",
    "grupo":   "Grupo / Origen",
    "fecha":   "Fecha",
    "ruta":    "Ruta",
    "estado":  "Estado",
    "ciudad":  "Ciudad",
    "factura": "Factura / N° Control",
    "reclamo": "Reclamo",
    "empresa": "Empresa",
    "rif":     "RIF",
    "error":   "Error",
}


def exportar_a_excel(df: pd.DataFrame, ruta_salida: Path) -> None:
    """
    Exporta el DataFrame a un archivo Excel (.xlsx) con formato profesional.

    Args:
        df:           DataFrame con los datos extraídos.
        ruta_salida:  Ruta completa del archivo de salida.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        raise ImportError(
            "openpyxl no está instalado. Ejecuta: pip install openpyxl"
        )

    from modules.config import CONFIG

    nombre_hoja = CONFIG.get("nombre_hoja", "Recolectas")
    color_cab   = CONFIG.get("color_cabecera", "FF1F3A5F")

    # ── Ordenar columnas ─────────────────────────────────────────────────────
    orden = CONFIG.get("campos_salida", list(df.columns))
    columnas_presentes = [c for c in orden if c in df.columns]
    df = df[columnas_presentes].rename(columns=COLUMNAS_LEGIBLES)

    # ── Crear workbook ───────────────────────────────────────────────────────
    wb = Workbook()

    # ── Hoja Principal ───────────────────────────────────────────────────────
    ws_datos = wb.active
    ws_datos.title = nombre_hoja

    # Estilos comunes
    fuente_cab = Font(
        name="Calibri", bold=True, color="FFFFFFFF", size=11
    )
    fill_cab = PatternFill(
        fill_type="solid", fgColor=color_cab
    )
    align_cab = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    border_delgado = Border(
        left=Side(style="thin", color="FFCCCCCC"),
        right=Side(style="thin", color="FFCCCCCC"),
        top=Side(style="thin", color="FFCCCCCC"),
        bottom=Side(style="thin", color="FFCCCCCC"),
    )
    fill_error = PatternFill(fill_type="solid", fgColor="FFFFCCCC")  # Rosa claro
    fill_impar = PatternFill(fill_type="solid", fgColor="FFF5F8FF")  # Azul muy suave
    align_celda = Alignment(vertical="center", wrap_text=False)

    # Escribir filas del DataFrame
    for fila_idx, fila in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws_datos.append(fila)
        fila_excel = fila_idx + 1  # 1-indexed

        if fila_idx == 0:
            # Cabecera
            for col_idx, _ in enumerate(fila, start=1):
                celda = ws_datos.cell(row=fila_excel, column=col_idx)
                celda.font = fuente_cab
                celda.fill = fill_cab
                celda.alignment = align_cab
                celda.border = border_delgado
        else:
            # Detectar si hay error en la fila
            col_error_nombre = COLUMNAS_LEGIBLES.get("error", "Error")
            hay_error = False
            if col_error_nombre in df.columns:
                col_pos = list(df.columns).index(col_error_nombre) + 1
                val_error = ws_datos.cell(row=fila_excel, column=col_pos).value
                hay_error = bool(val_error and str(val_error).strip())

            for col_idx, _ in enumerate(fila, start=1):
                celda = ws_datos.cell(row=fila_excel, column=col_idx)
                celda.alignment = align_celda
                celda.border = border_delgado
                if hay_error:
                    celda.fill = fill_error
                elif fila_idx % 2 == 0:
                    celda.fill = fill_impar

    # Ajustar ancho de columnas automáticamente
    for col_idx, col in enumerate(df.columns, start=1):
        letra = get_column_letter(col_idx)
        max_len = len(str(col))
        for row_idx in range(2, ws_datos.max_row + 1):
            val = ws_datos.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 60))
        ws_datos.column_dimensions[letra].width = max_len + 4

    # Altura de la cabecera
    ws_datos.row_dimensions[1].height = 30

    # Fijar la fila de cabecera (freeze panes)
    ws_datos.freeze_panes = "A2"

    # Filtros automáticos
    ws_datos.auto_filter.ref = ws_datos.dimensions

    # ── Hoja de Resumen ──────────────────────────────────────────────────────
    ws_res = wb.create_sheet(title="Resumen")
    col_error = COLUMNAS_LEGIBLES.get("error", "Error")

    total = len(df)
    exitosos = int((df[col_error] == "").sum()) if col_error in df.columns else total
    fallidos = total - exitosos

    # Calcular completitud por campo (excluir 'Archivo PDF' y 'Error')
    cols_datos = [
        c for c in df.columns
        if c not in (COLUMNAS_LEGIBLES.get("archivo", ""), col_error)
    ]
    completitud = {
        col: f"{int((df[col].astype(str).str.strip() != '').sum() / total * 100)}%"
        for col in cols_datos
        if total > 0
    }

    resumen_filas = [
        ["RESUMEN DE EXTRACCIÓN", ""],
        ["", ""],
        ["Total de PDFs procesados", total],
        ["Procesados con éxito",    exitosos],
        ["Con errores",             fallidos],
        ["", ""],
        ["COMPLETITUD POR CAMPO", ""],
    ] + [[col, pct] for col, pct in completitud.items()]

    fuente_titulo = Font(name="Calibri", bold=True, size=13, color="FF1F3A5F")
    fuente_sub    = Font(name="Calibri", bold=True, size=11)
    fuente_normal = Font(name="Calibri", size=11)

    for fila in resumen_filas:
        ws_res.append(fila)

    # Estilos del resumen
    ws_res["A1"].font = fuente_titulo
    ws_res["A7"].font = fuente_sub
    for row in ws_res.iter_rows(min_row=3, max_row=5, min_col=1, max_col=2):
        for cell in row:
            cell.font = fuente_normal

    ws_res.column_dimensions["A"].width = 32
    ws_res.column_dimensions["B"].width = 20

    # ── Guardar ──────────────────────────────────────────────────────────────
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta_salida))
    logger.info(f"Excel guardado: {ruta_salida.resolve()}")
    logger.info(f"  Hoja '{nombre_hoja}': {total} registros | {exitosos} exitosos | {fallidos} con error")
